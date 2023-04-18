import configuration_auto as conf1_auto
import openpyxl
import pandas as pd 
from converter_auto import *
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
import calendar
#import traceback

class Report_Generator:

    def __init__(self,mon,year):
        self.path1_save_lumen = conf1_auto.file_path.get('save_path_lumen')
        self.path1_save_brightspeed=conf1_auto.file_path.get('save_path_brightspeed')
        self.path2_extract_lumen =conf1_auto.file_path.get('extract_path_lumen')
        self.path2_extract_brightspeed =conf1_auto.file_path.get('extract_path_brightspeed')
        self.path3_template=conf1_auto.file_path.get('template_path')

        self.extract_filename= conf1_auto.file_name.get('extract_name')
        self.save_filename1= conf1_auto.file_name.get('change')
        self.save_filename2= conf1_auto.file_name.get('no_change')
        self.template_name= conf1_auto.file_name.get('template_name')
        self.year=year
        self.appli =conf1_auto.application.get('name')
        self.m, self.y,self.m1 ,self.day= Converter.mon_year(mon,year)
    def folder_check(self):
            try:
                self.file_path1_save=None
                self.file_path2_save=None

                if self.appli == 'lumen':
                    self.folder_path_extract = '{}/{}/{}/{}'.format(self.path2_extract_lumen,self.year,self.m1,self.extract_filename.format((self.appli).upper(),(self.m).upper(),self.day,self.year))
                    self.folder_path_save = '{}/{}/{}/'.format(self.path1_save_lumen,self.year,self.m1)
                    self.save_filename1=self.save_filename1.format((self.m).upper(),(self.y).upper(),(self.appli).upper())
                    self.save_filename2=self.save_filename2.format((self.m).upper(),(self.y).upper(),(self.appli).upper())
                elif self.appli == 'brightspeed':
                    self.folder_path_extract = '{}/{}/{}/{}'.format(self.path2_extract_brightspeed,self.year,self.m1,self.extract_filename.format((self.appli).upper(),(self.m).upper(),self.day,self.year))
                    self.folder_path_save = '{}/{}/{}/'.format(self.path1_save_brightspeed,self.year,self.m1)
                    self.save_filename1=self.save_filename1.format((self.m).upper(),(self.y).upper(),(self.appli).upper())
                    self.save_filename2=self.save_filename2.format((self.m).upper(),(self.y).upper(),(self.appli).upper())
                else:
                    raise ValueError("Invalid application name")
                    
                
                #FOR EXTRACTION PATH checking 
                if not os.path.isfile(self.folder_path_extract): 
                    raise FileNotFoundError("Extraction Folder {} not found or does'nt exists".format(self.extract_filename.format((self.appli).upper(),(self.m).upper(),self.day,self.year)))


                #To check save folder existes or not  and saving the file in that folder
                if not os.path.isdir(self.folder_path_save): 
                    os.makedirs(self.folder_path_save)
                
                files = os.listdir(self.folder_path_save)
                if  self.save_filename1  not in files and  self.save_filename2  not in files :
                    self.file_path1_save = os.path.join(self.folder_path_save, self.save_filename1)
                    self.file_path2_save = os.path.join(self.folder_path_save,self.save_filename2)
                else:
                    raise FileExistsError("Report '{}' and '{}' already exists for the given month and year".format(self.save_filename1,self.save_filename2))
            
            except (ValueError, FileNotFoundError, FileExistsError) as e:
                print(f"Exception occurred while checking the folder: {e}")

#########################################################################################
class Automation_billing(Report_Generator):
    def excel_report(self):
        try:
                super().folder_check()
                if self.file_path1_save is not None and self.file_path2_save is not None:
                    print("Creating files for both 'with changes' and 'with no changes'")
                    workbook=openpyxl.load_workbook(self.folder_path_extract)
                    worksheet=workbook['Summary - Daily View']

                    #FOR WITHCHANGE
                    filtered_rows_automation1=[]
                    filtered_rows_failuresplit1=[]

                    #FOR WITHOUT CHANGE
                    filtered_rows_automation2=[]
                    filtered_rows_failuresplit2=[]

                    month_num = list(calendar.month_abbr).index((self.m).capitalize())
                    num_days = calendar.monthrange(int(self.year), month_num)[1]#for geting no of days in the month
                    prow=''
                    for z,row in enumerate(worksheet.iter_rows(values_only=True,min_col=2,max_col=36,min_row=6,max_row=67),0):
                        if z<(num_days*2):
                            if row[0]!=None:
                                prow=row[0]
                            
                            row_appli=list(row[:1]+row[2:5])
                            if row[1]=='Change':#for with change report
                                row_appli[0]=prow
                                date_obj = datetime.strptime(row_appli[0],'%m-%d-%Y')
                                row_appli[0] = date_obj.strftime('%b/%d/%Y')
                                tot_fail=row[2]-row[3]
                                row_appli.insert(3,tot_fail)
                                if isinstance(row_appli[4], float):
                                    row_appli[4] = '{:.0%}'.format(row_appli[4])
                                filtered_rows_automation1.append(row_appli)
                                filtered_rows_failuresplit1.append([row_appli[0],(row[10]+row[20]+row[30]),(row[12]+row[22]+row[32]),(row[14]+row[24]+row[34]),tot_fail])

                            else:#for without change report
                                date_obj = datetime.strptime(row_appli[0],'%m-%d-%Y')
                                row_appli[0] = date_obj.strftime('%b/%d/%Y')
                                tot_fail=row[2]-row[3]
                                row_appli.insert(3,tot_fail)
                                if isinstance(row_appli[4], float):
                                    row_appli[4] = '{:.0%}'.format(row_appli[4])
                                filtered_rows_automation2.append(row_appli)
                                filtered_rows_failuresplit2.append([row_appli[0],(row[10]+row[20]+row[30]),(row[12]+row[22]+row[32]),(row[14]+row[24]+row[34]),tot_fail])
                        else:
                            break

                    #for with change        
                    df1 = pd.DataFrame(filtered_rows_automation1, columns =['Date','TotalSO#' ,'Success# ','Failure','Success%'])
                    df2 = pd.DataFrame(filtered_rows_failuresplit1, columns =['Date','Comm Error','Data Error','Gen Error','Total Failure'])
                    wb1= openpyxl.load_workbook(self.path3_template+ self.template_name)

                    ws1 = wb1.worksheets[0]
                    for r ,row  in enumerate(dataframe_to_rows(df1,index=False,header=False),4):
                        for c,val in enumerate(row,3):
                            ws1.cell(row=r,column=c,value=val)

                    ws2 = wb1.worksheets[1]
                    for r ,row  in enumerate(dataframe_to_rows(df2,index=False,header=False),4):
                        for c,val in enumerate(row,3):
                            ws2.cell(row=r,column=c,value=val)
                    wb1.save(self.file_path1_save)
                    

                    #for without change
                    df3 = pd.DataFrame(filtered_rows_automation2, columns =['Date','TotalSO#' ,'Success# ','Failure','Success%'])
                    df4 = pd.DataFrame(filtered_rows_failuresplit2, columns =['Date','Comm Error','Data Error','Gen Error','Total Failure'])
                    wb2= openpyxl.load_workbook(self.path3_template+ self.template_name)
                    ws3 = wb2.worksheets[0]

                    for r ,row  in enumerate(dataframe_to_rows(df3,index=False,header=False),4):
                        for c,val in enumerate(row,3):
                            ws3.cell(row=r,column=c,value=val)

                    ws4 = wb2.worksheets[1]
                    for r ,row  in enumerate(dataframe_to_rows(df4,index=False,header=False),4):
                        for c,val in enumerate(row,3):
                            ws4.cell(row=r,column=c,value=val)
                    wb2.save(self.file_path2_save)

                    if num_days != 31:
                        Converter.Excel_filter_Automation(self.file_path1_save)
                        Converter.Excel_filter_Automation(self.file_path2_save)
                    else:
                         print("No Empty rows in excel Sheets")
                    print("Report with 'with change' is done")
                    print("Report with 'without change' is done")

                    wb1.close()
                    wb2.close()
                    print("Excel report generated successfully.")

        except ValueError as ve:
            print(f"Exception in automation_report on Excelgenerator_auto: {ve}")
        except FileNotFoundError as fe:
            print(f"Exception in automation_report on Excelgenerator_auto: {fe}")
        except Exception as e:
            print(f"Exception in automation_report on Excelgenerator_auto: {e}")
            # tb = traceback.extract_tb(e.__traceback__)
            # filename, line_no, func, line = tb[-1]
            # err_msg = f"Error occurred in line {filename, line_no, func, line }: {e}"
            # print(err_msg)
