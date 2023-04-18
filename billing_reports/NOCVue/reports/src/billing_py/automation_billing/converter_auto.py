from datetime import datetime
from openpyxl.utils.cell import get_column_letter
import openpyxl
from calendar import monthrange
class Converter:
    
    @staticmethod
    def mon_year(mon, year):#FOR INPUT MONTH AND YEAR CONVERSION TO DESIRED FORMAT VALUE
      
        try:
            month_abbr = mon.upper()
            year_abbr = year[-2:]
            month_full = datetime.strptime(f"{mon}-{year_abbr}", "%b-%y").strftime("%B").lower()
            last_day = monthrange(int(year), datetime.strptime(mon, '%b').month)[1]
        except ValueError as e:
            print(f"Exception in Converter mon_year: {e}")
            return None, None, None,None
        else:
            return month_abbr, year_abbr, month_full,last_day
    
    @staticmethod
    def Excel_filter_Automation(file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            if wb==None:
                raise FileNotFoundError("Error in Excel_filter of Converter class: file_name or Sheetname not found")

            ws = wb.worksheets[1]#FOR FAILURE SPLITS CHECKING FOR EMPTY ROWS
            x=[]
            for row in ws.iter_rows(min_row=4, max_row=34, min_col=3, max_col=7):
                if row[0].value == None:
                    x.append(row[0].row)
                if row[0].value == 'Total':  
                    break
            
            ws1 = wb.worksheets[0]#FOR AUTOMATION CHECKING FOR EMPTY ROWS
            y=[]
            for row in ws1.iter_rows(min_row=4, max_row=34, min_col=3, max_col=7):
                if row[0].value == None:
                    y.append(row[0].row)
                if row[0].value == 'Total':
                    break
            
            
            
            #FOR FAILURE SPLITS reasigning the formulas
            start_row=4
            end_row=x[0]-1
            start_col=4
            end_col=7
            for i in reversed(x):
                ws.delete_rows(x[0],1)
            for col in range(start_col, end_col+1):

                start_cell = get_column_letter(col) + str(start_row)
                end_cell = get_column_letter(col) + str(end_row)
                sum_range = ":".join([start_cell, end_cell])


                total_col = ws.cell(row=end_row+1, column=col)
                sum_formula = f"=SUM({sum_range})"
                updated_formula = str(total_col.value).replace(str(total_col.value), sum_formula)
                total_col.value = updated_formula

            #FOR AUTOMATION reasigning the formulas
            start_row1=4
            end_row1=y[0]-1
            start_col1=4
            end_col1=7
            for i in reversed(y):
                ws1.delete_rows(y[0],1)
            for col in range(start_col1, start_col1+2):
                start_cell1 = get_column_letter(col) + str(start_row1)
                end_cell1 = get_column_letter(col) + str(end_row1)
                sum_range1 = ":".join([start_cell1, end_cell1])
                total_col1 = ws1.cell(row=end_row1+1, column=col)
                sum_formula1 = f"=SUM({sum_range1})"
                updated_formula1 = str(total_col1.value).replace(str(total_col1.value), sum_formula1)
                total_col1.value = updated_formula1

            total_Total_SO_cell1 = get_column_letter(start_col1) + str(end_row1+1)
            total_success_cell1 = get_column_letter(start_col1+1) + str(end_row1+1)

            #FOR total_row_failure COLUMN
            failure_formula = "-".join([total_Total_SO_cell1, total_success_cell1])
            total_row_failure = ws1.cell(row=end_row1+1, column=end_col1-1)
            failure_formula_total_col = f"={failure_formula}"
            updated_formula = str(total_row_failure.value).replace(str(total_row_failure.value), failure_formula_total_col)
            total_row_failure.value = updated_formula

            #FOR total_row_success COLUMN
            total_success_percentage_formula = "/".join([total_success_cell1,total_Total_SO_cell1])
            total_success_percentage=f"={total_success_percentage_formula}"
            total_row_success_percent = ws1.cell(row=end_row1+1, column=end_col1)
            updated_formula = str(total_row_success_percent.value).replace(str(total_row_success_percent.value), total_success_percentage)
            total_row_success_percent.value = updated_formula

            wb.save(file_path)
        except FileNotFoundError:
            print("File not found.")
        except Exception as e:
            print(f"Exception  in  Excel_filter_Automation: {e}")
        finally:
            wb.close()


