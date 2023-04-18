import configuration_CLI as conf1_CLI
import openpyxl
import pandas as pd 
from converter_CLI import *
from sql_data_extraction_CLI import *
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import calendar
#import traceback
class Report_Generator:

    def __init__(self,mon,year):

        self.path1_save_lumen = conf1_CLI.file_path.get('save_path_lumen')
        self.path1_save_brightspeed =conf1_CLI.file_path.get('save_path_brightspeed')
        self.path2_extract_lumen=conf1_CLI.file_path.get('extract_path_lumen')
        self.path2_extract_brightspeed=conf1_CLI.file_path.get('extract_path_brightspeed')
        self.path3_template =conf1_CLI.file_path.get('template_path')
        self.path4_sql= conf1_CLI.file_path.get('sql_path')

        self.extract_filename= conf1_CLI.file_name.get('extract_name')
        self.save_name= conf1_CLI.file_name.get('save_name')
        self.template_name= conf1_CLI.file_name.get('template_name')
        self.appli = conf1_CLI.application.get('name')
        #self.option = conf1_CLI.application.get('option')
        self.year=year
        self.m, self.y,self.m1 ,self.day= Converter.mon_year(mon,year)
        self.from_date,self.to_date = Converter.to_from(mon,year)
        if self.from_date==None:
            raise TypeError
        
    def folder_check(self,option):#type can be adtran or tellabs
        self.file_path_save=None
        self.file_path_extract=None
        self.option=option
        try:
            if self.appli == 'lumen':
                self.file_path_extract = '{}/{}/{}/{}'.format(self.path2_extract_lumen,self.year,self.m1,self.extract_filename.format((self.appli).upper(),(self.m).upper(),self.day,self.year))
                self.folder_path_save = '{}/{}/{}/'.format(self.path1_save_lumen,self.year,self.m1)
                self.save_filename=self.save_name.format(self.option,(self.m).upper(),(self.y).upper(),(self.appli).upper())
            elif self.appli == 'brightspeed':
                self.file_path_extract = '{}/{}/{}/{}'.format(self.path2_extract_brightspeed,self.year,self.m1,self.extract_filename.format((self.appli).upper(),(self.m).upper(),self.day,self.year))
                self.folder_path_save = '{}/{}/{}/'.format(self.path1_save_brightspeed,self.year,self.m1)
                self.save_filename=self.save_name.format(self.option,(self.m).upper(),(self.y).upper(),(self.appli).upper())
            else:
                raise ValueError("Invalid application name")
                
           
            #FOR EXTRACTION PATH checking 
            if not os.path.isfile(self.file_path_extract): 
                raise FileNotFoundError("Extraction Folder {} not found or does'nt exists".format(self.extract_filename.format((self.appli).upper(),(self.m).upper(),self.day,self.year)))

            #To check save folder existes or not  and saving the file in that folder
            if not os.path.isdir(self.folder_path_save): 
                os.makedirs(self.folder_path_save)

            if not os.path.isfile(os.path.join(self.folder_path_save, self.save_filename)): 
                self.file_path_save = os.path.join(self.folder_path_save, self.save_filename)
            else:
                raise FileExistsError("Report '{}'  already exists for the given month and year".format(self.save_filename))
        
        except (ValueError, FileNotFoundError, FileExistsError) as e:
            print(f"Exception occurred while checking the folder: {e}")

class CLI_Reports(Report_Generator):
    def excel_report(self,option):
        try:
            
            super().folder_check(option)
            if self.file_path_save != None:

                if self.option == 'Adtran':
                    print("Adtran File  Creation initiated")
                    access_sheet='CLI-Report Summary-Adtran'
                    save_sheet_name="Adtran-CLI"
                    database_option='{}-Lumen'.format((self.option).capitalize())
                    

                elif self.option=='Tellabs':
                    print("Tellabs File Creation initiated")
                    access_sheet='CLI-Report Summary-TELLABS'
                    save_sheet_name="Tellabs-CLI"
                    database_option=(self.option).capitalize()
                    
                else:
                    raise ValueError(f"{self.option} value is invalid ")

                workbook = openpyxl.load_workbook(self.file_path_extract )
                if workbook==None:
                    raise FileNotFoundError("Extraction File for {} Does'nt exist".format(self.appli))
                
                
                worksheet = workbook[access_sheet]
                filtered_rows = []

                for row in worksheet.iter_rows(values_only=True, min_col=2, max_col=7, min_row=5, max_row=35):
                    if row[0] is None:
                        break
                    filtered_rows.append(list(row))

                df = pd.DataFrame(filtered_rows, columns=[' ', 'TotalSO#', 'Comm Error', 'Success', 'Failure', 'Success %'])

                
                wb = openpyxl.load_workbook(self.path3_template +self.template_name)
                if wb==None:
                    raise Exception("Exception in CLI_Reports: template file not found")
                ws = wb.worksheets[0]
                ws.title=save_sheet_name
                
 

                for r, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=5):
                    for c, val in enumerate(row, start=3):
                        ws.cell(row=r, column=c, value=val)

                sql_object = SqlDataExtractor()
                raw_sql_data = sql_object.extract_sql_data(self.from_date,self.to_date,database_option,self.appli,self.path4_sql)
                result = [t[6:] for t in raw_sql_data]
                
                filtered_g = Converter.filter(result)
            
                ws1 = wb.worksheets[1]
                ws1.title="Failure-Split"
                df1 = pd.DataFrame(filtered_g, columns=['Date', 'Comm Error', 'Data Error', 'Gen Error', 'Total Failure'])

                for r1, row1 in enumerate(dataframe_to_rows(df1, index=False, header=False), start=4):
                    for c1, val1 in enumerate(row1, start=3):
                        ws1.cell(row=r1, column=c1, value=val1)

                wb.save(self.file_path_save)

                month_num = list(calendar.month_abbr).index((self.m).capitalize())
                num_days = calendar.monthrange(int(self.year), month_num)[1]
                if num_days!=31:
                    Converter.Excel_filter_CLI (self.file_path_save)
                else:
                    print("No Empty rows in excel Sheets")
                wb.close()
                print(f"{self.option} is saved")
        except Exception as e:
            (f"ExcelGenerator {self.option}_CLI Exception : {e}")
            # tb = traceback.extract_tb(e.__traceback__)
            # filename, line_no, func, line = tb[-1]
            # err_msg = f"Error occurred in line {filename, line_no, func, line }: {e}"
            # print(err_msg)
        

