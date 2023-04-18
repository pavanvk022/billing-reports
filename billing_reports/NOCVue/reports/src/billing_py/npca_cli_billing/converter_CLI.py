from datetime import datetime,timedelta
from openpyxl.utils.cell import get_column_letter
import openpyxl
from calendar import monthrange
class Converter:
    
    @staticmethod
    def mon_year(mon, year):#FOR INPUT MONTH AND YEAR CONVERSION TO DESIRED FORMAT VALUE
      
        try:
            month_abbr = mon.upper()
            year_abbr = year[-2:]
            month_full = datetime.strptime(f"{mon}-{year_abbr}", "%b-%y").strftime("%B").lower()
            last_day = monthrange(int(year), datetime.strptime(mon, '%b').month)[1]
        except ValueError as e:
            print(f"Exception in Converter mon_year: {e}")
            return None, None, None,None
        else:
            return month_abbr, year_abbr, month_full,last_day
        
    @staticmethod
    def to_from(mon, year):#TO GET 'TO_FROM' AND 'DATE_EXCLUDE' FOR SQL QUERY 
        try:
            # Construct the input date string using the input month and year
            date_string = f"01-{mon}-{year[-2:]}"
            date_obj = datetime.strptime(date_string, "%d-%b-%y")
            input_month = date_obj.month
            input_year = date_obj.year

            prev_month = input_month  if input_month > 1 else 12
            prev_year = input_year if input_month > 1 else input_year - 1
            first_day_month = datetime(prev_year, prev_month, 1)

            next_month = input_month + 1 if input_month < 12 else 1
            next_year = input_year if input_month < 12 else input_year + 1
            first_day_next_month = datetime(next_year, next_month, 1)

            to_from = first_day_month.strftime("%d-%m-%y")
            date_exclude = first_day_next_month.strftime("%d-%m-%y")
            return to_from, date_exclude

        except ValueError as ve:
            print("Exception in Converter to_from:", ve)
            return None, None

    @staticmethod
    def filter(data):#FOR ADDING MISSING DATES AND APPENDING THE VALUES WITH ZERO VALUES
        dates = []
        for date_str, *_ in data:
            try:
                date = datetime.strptime(date_str, '%b/%d/%Y')
                dates.append(date)
            except ValueError:
                pass

        if dates:
            start_date = min(dates)
            end_date = max(dates)
            delta = timedelta(days=1)

            all_dates = []
            while start_date <= end_date:
                all_dates.append(start_date.strftime('%b/%d/%Y'))
                start_date += delta
            all_dates=[i.upper() for i in all_dates]
            new_data = []
            for d in all_dates:
                for date_str, *values in data:
                    if d == date_str:
                        new_data.append((date_str, *values))
                        break
                else:
                    new_data.append((d, 0, 0, 0, 0))

            return new_data

        else:
            return data
        
    @staticmethod
    def Excel_filter_CLI(file_path):
        try:
            
            wb = openpyxl.load_workbook(file_path)
            
            if wb==None:
                raise Exception("Error in Excel_filter of Converter class: file_name or Sheetname not found")

            #FOR FAILURE SPLIT CHECKING FOR EMPTY ROWS
            ws = wb.worksheets[1]
            x=[]
            for row in ws.iter_rows(min_row=4, max_row=34, min_col=3, max_col=7):
                if row[0].value == None:
                    x.append(row[0].row)
                if row[0].value == 'Total':  
                    break
            
            #FOR CLI REPORT CHECKING FOR EMPTY ROWS
            ws1 = wb.worksheets[0]
            y=[]
            for row in ws1.iter_rows(min_row=5, max_row=35, min_col=3, max_col=8):
                if row[0].value == None:
                    y.append(row[0].row)
                if row[0].value == 'Total':
                    break
            
            
            #FOR FAILURE SPLIT REPORT FORMULA REASSIGNMENT
            start_row=4
            end_row=x[0]-1
            start_col=4
            end_col=7
            for i in reversed(x):
                ws.delete_rows(x[0],1)
            for col in range(start_col, end_col+1):

                start_cell = get_column_letter(col) + str(start_row)
                end_cell = get_column_letter(col) + str(end_row)
                sum_range = ":".join([start_cell, end_cell])


                total_col = ws.cell(row=end_row+1, column=col)
                sum_formula = f"=SUM({sum_range})"
                updated_formula = str(total_col.value).replace(str(total_col.value), sum_formula)
                total_col.value = updated_formula

            #FOR CLI REPORT FORMULA REASSIGNMENT
            start_row1=5
            end_row1=y[0]-1
            start_col1=4
            end_col1=8
            for i in reversed(y):
                ws1.delete_rows(y[0],1)
            for col in range(start_col1, end_col1):
                start_cell1 = get_column_letter(col) + str(start_row1)
                end_cell1 = get_column_letter(col) + str(end_row1)
                sum_range1 = ":".join([start_cell1, end_cell1])
                total_col1 = ws1.cell(row=end_row1+1, column=col)
                sum_formula1 = f"=SUM({sum_range1})"
                updated_formula1 = str(total_col1.value).replace(str(total_col1.value), sum_formula1)
                total_col1.value = updated_formula1

            suceess_total_cell = get_column_letter(end_col1-2) + str(end_row1+1)
            commerror_total_cell = get_column_letter(end_col1-3) + str(end_row1+1)
          
            #FOR SUCCESS PERCENTAGE COLUMN 
            sucess_percent_formula = "/".join( [suceess_total_cell,commerror_total_cell])
            sucess_percent_formula_col = f"={sucess_percent_formula}"
            total_row = ws1.cell(row=end_row1+1, column=end_col1)
            updated_formula = str(total_row.value).replace(str(total_row.value),  sucess_percent_formula_col)
            total_row.value = updated_formula

            wb.save(file_path)
            wb.close()
        except FileNotFoundError:
            print("File not found.")
        except Exception as e:
            print(f"Exception opening file in Excel Converter: {e}")
        
  